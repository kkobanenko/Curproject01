from __future__ import annotations
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
import hashlib

from .document_parser import DocumentParser

logger = logging.getLogger(__name__)

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not available, XLSX parsing disabled")

class XLSXParser(DocumentParser):
    """Парсер для XLSX файлов с поддержкой таблиц и метаданных"""
    
    def __init__(self):
        super().__init__()
        self.supported_mime_types = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ]
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Проверяет доступность зависимостей"""
        if not XLSX_AVAILABLE:
            logger.error("openpyxl not available - XLSX parsing will fail")
    
    def can_parse(self, file_path: Path) -> bool:
        """Проверяет, может ли парсер обработать XLSX файл"""
        return (file_path.suffix.lower() == '.xlsx' and 
                file_path.exists() and 
                XLSX_AVAILABLE)
    
    def parse(self, file_path: Path) -> Dict[str, Any]:
        """Парсит XLSX документ"""
        if not self.can_parse(file_path):
            raise ValueError(f"Cannot parse file: {file_path}")
        
        # Вычисляем SHA256
        sha256 = self._calculate_sha256(file_path)
        
        # Определяем MIME тип
        mime_type = self.get_mime_type(file_path)
        
        # Парсим содержимое
        content = self._extract_content(file_path)
        
        return {
            'title': content.get('title'),
            'text': content.get('text', ''),
            'tables': content.get('tables', []),
            'pages': content.get('pages', []),
            'sha256': sha256,
            'mime_type': mime_type,
            'needs_ocr': False,  # XLSX не нужен OCR
            'metadata': content.get('metadata', {}),
            'sheet_count': content.get('sheet_count', 0),
            'table_count': len(content.get('tables', [])),
            'has_tables': len(content.get('tables', [])) > 0
        }
    
    def _extract_content(self, file_path: Path) -> Dict[str, Any]:
        """Извлекает содержимое XLSX"""
        content = {
            'text': '',
            'tables': [],
            'pages': [],
            'metadata': {},
            'sheet_count': 0
        }
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Метаданные
            if workbook.properties:
                content['metadata'] = {
                    'title': workbook.properties.title,
                    'author': workbook.properties.creator,
                    'subject': workbook.properties.subject,
                    'created': workbook.properties.created,
                    'modified': workbook.properties.modified,
                    'last_modified_by': workbook.properties.lastModifiedBy,
                    'category': workbook.properties.category,
                    'keywords': workbook.properties.keywords,
                    'comments': workbook.properties.description,
                    'company': workbook.properties.company,
                    'manager': workbook.properties.manager
                }
                content['title'] = content['metadata'].get('title')
            
            content['sheet_count'] = len(workbook.sheetnames)
            
            # Обрабатываем каждый лист
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._extract_sheet_data(sheet, sheet_name)
                
                if sheet_data:
                    content['tables'].append(sheet_data)
                    content['text'] += f"Sheet: {sheet_name}\n"
                    content['text'] += self._sheet_to_text(sheet_data)
                    content['text'] += "\n\n"
            
            # Создаем "страницы" для совместимости
            content['pages'] = [{'page_num': i+1, 'text': f"Sheet {i+1}: {name}"} 
                              for i, name in enumerate(workbook.sheetnames)]
            
        except Exception as e:
            logger.error(f"Error parsing XLSX {file_path}: {e}")
            raise
        
        return content
    
    def _extract_sheet_data(self, sheet, sheet_name: str) -> Optional[Dict[str, Any]]:
        """Извлекает данные из листа"""
        # Получаем границы данных
        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column
        
        if min_row is None or max_row is None:
            return None
        
        # Извлекаем данные
        sheet_data = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value
                
                # Обрабатываем различные типы данных
                if cell_value is None:
                    cell_value = ""
                elif isinstance(cell_value, (int, float)):
                    cell_value = str(cell_value)
                elif isinstance(cell_value, str):
                    cell_value = cell_value.strip()
                else:
                    cell_value = str(cell_value)
                
                row_data.append(cell_value)
            
            # Пропускаем пустые строки
            if any(cell.strip() for cell in row_data if cell):
                sheet_data.append(row_data)
        
        if not sheet_data:
            return None
        
        # Анализируем структуру
        structure = self._analyze_sheet_structure(sheet_data)
        
        return {
            'sheet_name': sheet_name,
            'data': sheet_data,
            'rows': len(sheet_data),
            'cols': len(sheet_data[0]) if sheet_data else 0,
            'structure': structure,
            'table_index': len(sheet_data)  # Временный индекс
        }
    
    def _analyze_sheet_structure(self, sheet_data: List[List[str]]) -> Dict[str, Any]:
        """Анализирует структуру листа"""
        if not sheet_data:
            return {}
        
        rows = len(sheet_data)
        cols = len(sheet_data[0]) if sheet_data else 0
        
        # Определяем заголовки (первая строка)
        headers = sheet_data[0] if sheet_data else []
        
        # Анализируем типы данных в колонках
        column_types = []
        column_stats = []
        
        for col_idx in range(cols):
            col_data = [row[col_idx] for row in sheet_data[1:] if col_idx < len(row)]
            col_type = self._infer_column_type(col_data)
            col_stats = self._analyze_column_stats(col_data)
            
            column_types.append(col_type)
            column_stats.append(col_stats)
        
        return {
            'headers': headers,
            'column_types': column_types,
            'column_stats': column_stats,
            'has_merged_cells': False,  # openpyxl не поддерживает объединение ячеек
            'is_empty': all(all(not cell for cell in row) for row in sheet_data),
            'data_types': {
                'numeric_columns': [i for i, t in enumerate(column_types) if t == 'numeric'],
                'date_columns': [i for i, t in enumerate(column_types) if t == 'date'],
                'text_columns': [i for i, t in enumerate(column_types) if t == 'text']
            }
        }
    
    def _infer_column_type(self, column_data: List[str]) -> str:
        """Определяет тип данных в колонке"""
        if not column_data:
            return 'empty'
        
        # Подсчитываем типы
        numeric_count = 0
        date_count = 0
        empty_count = 0
        
        for value in column_data:
            if not value or value.strip() == '':
                empty_count += 1
                continue
            
            # Проверяем на число
            try:
                float(value.replace(',', '').replace(' ', '').replace('$', '').replace('%', ''))
                numeric_count += 1
            except ValueError:
                pass
            
            # Простая проверка на дату
            if any(separator in value for separator in ['/', '-', '.', ':', ' ']):
                if len(value.split()) >= 2:  # Должно быть минимум 2 части
                    date_count += 1
        
        total_non_empty = len(column_data) - empty_count
        
        if total_non_empty == 0:
            return 'empty'
        elif numeric_count / total_non_empty > 0.7:
            return 'numeric'
        elif date_count / total_non_empty > 0.7:
            return 'date'
        else:
            return 'text'
    
    def _analyze_column_stats(self, column_data: List[str]) -> Dict[str, Any]:
        """Анализирует статистику колонки"""
        non_empty_data = [v for v in column_data if v and v.strip()]
        
        if not non_empty_data:
            return {'count': 0, 'unique': 0, 'min': None, 'max': None}
        
        unique_values = set(non_empty_data)
        
        # Пытаемся найти числовые значения
        numeric_values = []
        for value in non_empty_data:
            try:
                numeric_values.append(float(value.replace(',', '').replace(' ', '')))
            except ValueError:
                pass
        
        stats = {
            'count': len(non_empty_data),
            'unique': len(unique_values),
            'min': min(non_empty_data) if non_empty_data else None,
            'max': max(non_empty_data) if non_empty_data else None
        }
        
        if numeric_values:
            stats.update({
                'numeric_count': len(numeric_values),
                'numeric_min': min(numeric_values),
                'numeric_max': max(numeric_values),
                'numeric_avg': sum(numeric_values) / len(numeric_values)
            })
        
        return stats
    
    def _sheet_to_text(self, sheet_data: Dict[str, Any]) -> str:
        """Конвертирует лист в текстовое представление"""
        data = sheet_data.get('data', [])
        if not data:
            return "Empty sheet"
        
        text_lines = []
        
        # Добавляем заголовки
        if data:
            text_lines.append(" | ".join(str(cell) for cell in data[0]))
            text_lines.append("-" * len(text_lines[0]))
            
            # Добавляем данные
            for row in data[1:]:
                text_lines.append(" | ".join(str(cell) for cell in row))
        
        return "\n".join(text_lines)
    
    def _calculate_sha256(self, file_path: Path) -> str:
        """Вычисляет SHA256 хеш файла"""
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256_hash.update(chunk)
        return sha256_hash.hexdigest()
